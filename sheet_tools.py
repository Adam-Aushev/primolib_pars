from openpyxl import load_workbook
from openpyxl import Workbook
from pathlib import Path
import os 


# этот кусок делает запись в документ эксель, нужно название документа и инфа в лист формате
def write_data(sheet_name, list_info: list): # нужно название с расширением .xlsx
    sheet_dir = 'sheet_dir'
    os.mkdir(sheet_dir) if not os.path.isdir(sheet_dir) else 2
    sheet_name = os.path.splitext(os.path.basename(sheet_name))[0]
    sheet_path = os.path.join(sheet_dir, sheet_name)
    sheet_path = f'{sheet_path}.xlsx'
    workbook = Workbook()
    workbook.save(sheet_path)
    workbook = load_workbook(sheet_path)
    page = workbook.active
    headers = list(list_info[0])
    page.append(headers) # type: ignore

    for x in list_info:
        page.append(list(x.values())) # type: ignore


    workbook.save(sheet_path)
    return sheet_path
