import requests
from bs4 import BeautifulSoup
import time
import os
import random
import sys 
import json
import re
from openpyxl import load_workbook
import web_attrib
from tkinter import *
from tkinter import ttk
from tkinter import filedialog

def save_file(file):
    with open('ris_file.gif', 'wb') as ris:
        ris.write(file)


def get_proxy():
    url='http://htmlweb.ru/json/proxy/get?api_key=befa1738418a4056c2d4b0eeddb08165'
    if not os.path.isfile('proxy_file.txt'):
        respose = requests.get(url)
        proxy_dict = respose.text
        with open('proxy_file.txt', 'w', encoding='utf-8') as proxy_file:
            proxy_file.write(proxy_dict)
    with open('proxy_file.txt', 'r', encoding='utf-8') as proxy_file:
        proxy_dict = json.loads(proxy_file.read())
    proxy_list = []
    for i in range(20):
        proxy_list.append({proxy_dict[str(i)]['type'].lower().split(',')[-1]:proxy_dict[str(i)]['name']})
    return proxy_list


def save_site(url, site_file, headers, params, cookies):
    for each_iter in range(10):
        source_folder = 'source/'
        site_file = os.path.join(source_folder, site_file)
        site_file = site_file if '.html' in site_file else f'{site_file}.html'
        if 'test' in ''.join(sys.argv):
            print('-----------test_mode-------------')
            if not os.path.isdir(source_folder):
                os.mkdir(source_folder)
            if not os.path.isfile(site_file):
                req = requests.get(url = url, headers=headers, params=params, cookies=cookies)
                with open(site_file, 'w', encoding='utf-8') as file:
                    file.write(req.text)
            with open(site_file, 'r', encoding='utf-8') as file:
                src = file.read()
            soup = BeautifulSoup(src, 'lxml')
        else:
            time.sleep(random.randint(1, 2))
            src = requests.get(url = url, headers=headers, params=params, cookies=cookies)
            src = src.text
            soup = BeautifulSoup(src, 'lxml')
        if not soup.text.strip():
            print('soup error', each_iter)
            continue
        elif 'если проблема сохранится, пожалуйста, обратитесь к' in src:
            print('error если проблема сохранится')
            continue
        break
    return soup


def get_sheet(sheet_name, sheet_row='', status='Найдено'):
    workbook = load_workbook(sheet_name)
    sheet = workbook.active
    sheet_list = []
    for each_row in range(2, sheet.max_row+1):
        title = ''
        year = ''
        author = ''
        if any([cell.value is not None for cell in sheet[each_row]]):
            if sheet.cell(column=1, row=each_row).value != None:
                title = sheet.cell(column=1, row=each_row).value
            if sheet.cell(column=2, row=each_row).value != None:
                year = sheet.cell(column=2, row=each_row).value
            if sheet.cell(column=3, row=each_row).value != None:
                author = sheet.cell(column=3, row=each_row).value
            sheet_list.append({'title':title,
                                'year':year,
                                'author':author})
    if sheet_row:
        sheet.cell(row=1, column=4, value='Status')
        sheet.cell(row=sheet_row, column=4, value=status)
        workbook.save(sheet_name)
    return sheet_list






def product(xlsx_files, export_dir):
    headers = web_attrib.headers
    cookies = ''#web_attrib.cookies
    params = web_attrib.params
    with open('export_path.txt', 'r', encoding='utf-8') as file:
        export_dir = file.read()
    for xlsx in xlsx_files:
        xlsx = xlsx.strip()
        sheet_list = get_sheet(xlsx)
        mark_folder = f"{export_dir}/{os.path.splitext(os.path.split(xlsx)[1])[0]}_rusmark_files"
        print('---------',mark_folder)
        os.mkdir(mark_folder) if not os.path.isdir(mark_folder) else 1
        for each_row in sheet_list:
            params['vl(freeText0)'] = each_row['author']
            params['vl(freeText1)'] = each_row['title']
            params['vl(freeText2)'] = each_row['year']
            cicle_count = 1 + len(each_row['title'].split())
            url = 'https://primo.nlr.ru/primo_library/libweb/action/search.do'
            search_page = f"search_page_{re.sub('[:,. ]', '_', each_row['author'])}_{len(each_row['title'])}.html"
            for search_iter in range(cicle_count):
                print(params['vl(freeText0)'], params['vl(freeText1)'], params['vl(freeText2)'])
                soup = save_site(url, search_page, headers=headers, params=params, cookies=cookies)
                result_count = soup.find(id='resultsNumbersTile')
                result_count = result_count.text.strip().split('\n')[0]
                result_count = re.sub('[^0-9]', ' ', result_count).strip().split()
                result_count = int(result_count[-1])
                print('result_count -', result_count)
                if result_count > 0:
                    cards = soup.find(class_='EXLResultsTable')
                    cards = cards.find_all(class_=re.compile('EXLResultMediaTYPE'))
                    card_list = []
                    for each_card in cards:
                        link = each_card.find(class_='EXLResultTitle')
                        name = link.find('a').text
                        link = link.find('a').get('href')
                        media_type = each_card.find(class_='EXLThumbnailCaption').text
                        card_list.append({'name':name, 'link':link, 'nedia_type':media_type, 'search_link':''})
                    break
                else:
                    if len(params['vl(freeText0)'].split()) > 1:
                        params['vl(freeText0)'] = params['vl(freeText0)'].split()[0]
                    elif len(params['vl(freeText1)'].split()) > 1:    
                        params['vl(freeText1)'] = ' '.join(params['vl(freeText1)'].split()[:-1])

            for each_dict in card_list[:1]:
                file_name =  re.sub('[:,.()\\ ]', '_', each_dict['name'])
                each_link = f"https://primo.nlr.ru/primo_library/libweb/action/{each_dict['link']}"
                soup = save_site(f'{each_link}.html', file_name, headers=headers, params=params, cookies=cookies)
                rusmarc = soup.find(string=re.compile('RUSMARC ISO2709'))
                rusmarc = rusmarc.parent.get('href')
                rusmarc = requests.get(rusmarc).content
                status = 'Найдено'
                if soup.find(string='Автор:'):
                    author = soup.find(string='Автор:').parent.parent
                    author = author.find('a').text
                    author = re.sub('[:,.() ]', '_', author).replace('\\', '')
                    print('file name -----------', file_name, '\n author ----------------', author)
                    file_name = f"{author}_{file_name}"
                with open(os.path.join(mark_folder, f'{file_name}.mrc'), 'wb') as mark_file:
                    mark_file.write(rusmarc)
                if len(card_list) > 1:
                    status = f'{len(card_list)} Найдено'
                get_sheet(xlsx, sheet_row=sheet_list.index(each_row)+2, status=status)
                

if __name__ == '__main__':
    xlsx_files = os.listdir()
    xlsx_files = [ each for each in xlsx_files if '.xlsx' in each ]
    product(xlsx_files, './')
    print('Процесс завершен!')